#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import json
from datetime import datetime as dt
import openpyxl

# Load and aggregate data
with open(r"C:\Users\Michael\Documents\Tendermint\Gno.land\GNOT Distribution Sim\snapshot_consolidated_10562840.json", "r") as file:
    data = json.load(file)

votes_df = pd.read_csv(r"C:\Users\Michael\Documents\Tendermint\Gno.land\GNOT Distribution Sim\votes-unique.csv")
votes_df.columns = ["timestamp", "address", "vote"]

# List of addresses to omit
exclude_addresses = ['cosmos1fl48vsnmsdzcv85q5d2q4z5ajdha8yu34mf0eh', 'cosmos1tygms3xhhs3yv487phx3dw4a95jn7t7lpm470r']

# Filter from votes DataFrame
votes_df = votes_df[~votes_df['address'].isin(exclude_addresses)]

# Filter from data list
data = [entry for entry in data if entry['address'] not in exclude_addresses]

# Prop 69
multipliers = {
    '': 0.7,  # Abstain
    'VOTE_OPTION_NO': 1.1,
    'VOTE_OPTION_NO_WITH_VETO': 1.3,
    'VOTE_OPTION_YES': 0
}

# Adjust ATOM based on multiplier
adjusted_data = []
votes_dict = votes_df.set_index('address')['vote'].to_dict()
for entry in data:
    address = entry['address']
    uatom = sum(float(coin['amount']) for coin in entry['coins'] if coin['denom'] == 'uatom')
    atom = uatom / 1e6
    vote = votes_dict.get(address, '')
    adjusted_atom = atom * multipliers.get(vote, 1)
    adjusted_data.append({
        'address': address,
        'atom': atom,
        'adjusted_atom': adjusted_atom,
        'vote': vote
    })

adjusted_df = pd.DataFrame(adjusted_data)
total_atom = adjusted_df['atom'].sum()  # Manually adjust later based on explorer/snapshot
adjusted_df['atom_percentage'] = (adjusted_df['atom'] / total_atom)
total_adjusted_atom = adjusted_df['adjusted_atom'].sum()
adjusted_df['adjusted_atom_percentage'] = (adjusted_df['adjusted_atom'] / total_adjusted_atom)
adjusted_df['gnot_percentage'] = adjusted_df['adjusted_atom_percentage']

# Deduct New Tendermint's allocation from other addresses
new_tendermint_percentage = 0.25
adjusted_df['gnot_percentage'] *= (1 - new_tendermint_percentage)

# Insert New Tendermint data
new_tendermint_data = {
    'address': 'New Tendermint',
    'atom': 0,
    'adjusted_atom': 0,
    'vote': '',
    'atom_percentage': 0,
    'adjusted_atom_percentage': 0,
    'gnot_percentage': new_tendermint_percentage
}
adjusted_df = adjusted_df.append(new_tendermint_data, ignore_index=True)

# Airdrop options
def option_1(dataframe, genesis_mint=750e6):
    dataframe['gnot_allocated'] = dataframe['gnot_percentage'] * genesis_mint
    return dataframe

def option_2(dataframe, total_gnot=1e9):
    dataframe['gnot_allocated'] = dataframe['gnot_percentage'] * total_gnot
    return dataframe

def option_3(dataframe, total_gnot=1e9, sub_dao_amount=250e6):
    dataframe['gnot_allocated'] = dataframe['gnot_percentage'] * (total_gnot - sub_dao_amount)
    return dataframe

def option_4(dataframe, total_gnot=1e9):
    dataframe['gnot_allocated'] = dataframe['gnot_percentage'] * total_gnot
    return dataframe

option_1_df = option_1(adjusted_df.copy())
option_2_df = option_2(adjusted_df.copy())
option_3_df = option_3(adjusted_df.copy())
option_4_df = option_4(adjusted_df.copy())

# Date
now = dt.now()
formatted_date = now.strftime("%m-%d-%Y_%H_%M")

# Export
output_path = "C:\\Users\\Michael\\Documents\\Tendermint\\Gno.land\\GNOT Distribution Sim\\omit_airdrop_sim_" + formatted_date + ".xlsx"
with pd.ExcelWriter(output_path) as writer:
    adjusted_df.to_excel(writer, sheet_name="Base Data", index=False)
    option_1_df.to_excel(writer, sheet_name="Option 1", index=False)
    option_2_df.to_excel(writer, sheet_name="Option 2", index=False)
    option_3_df.to_excel(writer, sheet_name="Option 3", index=False)
    option_4_df.to_excel(writer, sheet_name="Option 4", index=False)

# openpyxl Excel edits and column formatting
wb = openpyxl.load_workbook(output_path)
for sheetname in ['Base Data', 'Option 1', 'Option 2', 'Option 3', 'Option 4']:
    ws = wb[sheetname]
    # Percentage columns are columns E, F, and G
    for col in ['E', 'F', 'G']:
        for row in range(2, ws.max_row + 1):  # Start from 2 to skip header
            ws[f"{col}{row}"].number_format = '0.00%'

wb.save(output_path)
wb.save(output_path)

